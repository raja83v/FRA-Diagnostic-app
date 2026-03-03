"""
Reports router — PDF report generation and Excel export.

Endpoints:
  GET /api/v1/reports/generate/{transformer_id}  → PDF analysis report
  GET /api/v1/reports/export/measurements         → Excel export of measurements
  GET /api/v1/reports/export/analyses              → Excel export of analyses
"""
import io
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.transformer import Transformer
from app.models.measurement import FRAMeasurement
from app.models.fault_analysis import FaultAnalysis, FaultType
from app.models.recommendation import Recommendation

router = APIRouter(prefix="/api/v1/reports", tags=["Reports"])


# ── PDF Report ──────────────────────────────────────────────────────────────

def _build_pdf(transformer: Transformer, measurements, analyses, recommendations) -> bytes:
    """Build a professional PDF analysis report for a transformer."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm, cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        title=f"FRA Report – {transformer.name}",
        author="FRA Diagnostic Software",
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        "SectionHeading",
        parent=styles["Heading2"],
        textColor=colors.HexColor("#0D9488"),
        spaceAfter=6,
    ))
    styles.add(ParagraphStyle(
        "SmallGrey",
        parent=styles["Normal"],
        fontSize=8,
        textColor=colors.grey,
    ))
    teal = colors.HexColor("#0D9488")

    story: list = []

    # ── Title block ──
    story.append(Paragraph("FRA Diagnostic Report", styles["Title"]))
    story.append(Spacer(1, 4 * mm))
    story.append(HRFlowable(width="100%", thickness=1, color=teal))
    story.append(Spacer(1, 4 * mm))

    # ── Transformer Info ──
    story.append(Paragraph("Transformer Details", styles["SectionHeading"]))
    info_data = [
        ["Name", transformer.name],
        ["Serial Number", transformer.serial_number or "—"],
        ["Manufacturer", transformer.manufacturer or "—"],
        ["Substation", transformer.substation or "—"],
        ["Voltage Rating", f"{transformer.voltage_rating_kv} kV" if transformer.voltage_rating_kv else "—"],
        ["Power Rating", f"{transformer.power_rating_mva} MVA" if transformer.power_rating_mva else "—"],
        ["Criticality", (transformer.criticality or "standard").capitalize()],
        ["Total Measurements", str(len(measurements))],
    ]
    info_table = Table(info_data, colWidths=[4.5 * cm, 12 * cm])
    info_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#334155")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("LINEBELOW", (0, 0), (-1, -2), 0.5, colors.HexColor("#E2E8F0")),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 6 * mm))

    # ── Analysis Results ──
    if analyses:
        story.append(Paragraph("ML Analysis Results", styles["SectionHeading"]))
        header = ["Date", "Fault Type", "Probability", "Confidence", "Health Score", "Model"]
        rows = [header]
        for a in analyses:
            rows.append([
                a.created_at.strftime("%Y-%m-%d %H:%M") if a.created_at else "—",
                (a.fault_type or "").replace("_", " ").title(),
                f"{a.probability_score:.1%}" if a.probability_score else "—",
                f"{a.confidence_level:.1%}" if a.confidence_level else "—",
                f"{a.health_score:.0f}" if a.health_score is not None else "—",
                a.model_version or "—",
            ])
        analysis_table = Table(rows, repeatRows=1)
        analysis_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F0FDFA")),
            ("TEXTCOLOR", (0, 0), (-1, 0), teal),
            ("LINEBELOW", (0, 0), (-1, 0), 1, teal),
            ("LINEBELOW", (0, 1), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("ALIGN", (2, 0), (-1, -1), "CENTER"),
        ]))
        story.append(analysis_table)
        story.append(Spacer(1, 6 * mm))

        # Fault probability breakdown for latest analysis
        latest = analyses[0]
        if latest.all_probabilities:
            story.append(Paragraph("Fault Probability Breakdown (Latest)", styles["SectionHeading"]))
            prob_rows = [["Fault Type", "Probability"]]
            for ft, prob in sorted(latest.all_probabilities.items(), key=lambda x: -x[1]):
                prob_rows.append([ft.replace("_", " ").title(), f"{prob:.2%}"])
            prob_table = Table(prob_rows, colWidths=[8 * cm, 4 * cm])
            prob_table.setStyle(TableStyle([
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F0FDFA")),
                ("TEXTCOLOR", (0, 0), (-1, 0), teal),
                ("LINEBELOW", (0, 0), (-1, 0), 1, teal),
                ("LINEBELOW", (0, 1), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("ALIGN", (1, 0), (1, -1), "CENTER"),
            ]))
            story.append(prob_table)
            story.append(Spacer(1, 6 * mm))
    else:
        story.append(Paragraph("No ML analyses have been run for this transformer yet.", styles["Normal"]))
        story.append(Spacer(1, 6 * mm))

    # ── Recommendations ──
    if recommendations:
        story.append(Paragraph("Maintenance Recommendations", styles["SectionHeading"]))
        rec_header = ["Urgency", "Status", "Action", "Due Date"]
        rec_rows = [rec_header]
        for r in recommendations:
            rec_rows.append([
                (r.urgency or "").replace("_", " ").title(),
                (r.status or "").replace("_", " ").title(),
                r.action_description or "—",
                r.due_date.strftime("%Y-%m-%d") if r.due_date else "—",
            ])
        rec_table = Table(rec_rows, colWidths=[2.5 * cm, 2.5 * cm, 8.5 * cm, 3 * cm], repeatRows=1)
        rec_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FFF7ED")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#C2410C")),
            ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#C2410C")),
            ("LINEBELOW", (0, 1), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(rec_table)
        story.append(Spacer(1, 6 * mm))

    # ── Measurements Table ──
    if measurements:
        story.append(Paragraph("Measurement History", styles["SectionHeading"]))
        m_header = ["Date", "Winding Config", "Vendor", "Freq Range", "Data Points"]
        m_rows = [m_header]
        for m in measurements[:20]:  # limit for page space
            freq_min = min(m.frequency_hz) if m.frequency_hz else 0
            freq_max = max(m.frequency_hz) if m.frequency_hz else 0
            m_rows.append([
                m.measurement_date.strftime("%Y-%m-%d") if m.measurement_date else "—",
                m.winding_config or "—",
                m.vendor or "—",
                f"{freq_min/1000:.0f}–{freq_max/1000:.0f} kHz",
                str(len(m.frequency_hz)) if m.frequency_hz else "0",
            ])
        m_table = Table(m_rows, repeatRows=1)
        m_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F0FDFA")),
            ("TEXTCOLOR", (0, 0), (-1, 0), teal),
            ("LINEBELOW", (0, 0), (-1, 0), 1, teal),
            ("LINEBELOW", (0, 1), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("ALIGN", (3, 0), (-1, -1), "CENTER"),
        ]))
        story.append(m_table)
        story.append(Spacer(1, 6 * mm))

    # ── Footer ──
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#CBD5E1")))
    story.append(Spacer(1, 2 * mm))
    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    story.append(Paragraph(
        f"Generated by FRA Diagnostic Software on {now_str}",
        styles["SmallGrey"],
    ))

    doc.build(story)
    return buf.getvalue()


@router.get("/generate/{transformer_id}")
def generate_pdf_report(transformer_id: str, db: Session = Depends(get_db)):
    """Generate a comprehensive PDF analysis report for a transformer."""
    transformer = db.query(Transformer).filter(Transformer.id == transformer_id).first()
    if not transformer:
        raise HTTPException(404, "Transformer not found")

    measurements = (
        db.query(FRAMeasurement)
        .filter(FRAMeasurement.transformer_id == transformer_id)
        .order_by(FRAMeasurement.measurement_date.desc())
        .all()
    )

    measurement_ids = [m.id for m in measurements]
    analyses = (
        db.query(FaultAnalysis)
        .filter(FaultAnalysis.measurement_id.in_(measurement_ids))
        .order_by(FaultAnalysis.created_at.desc())
        .all()
    ) if measurement_ids else []

    recommendations_list = (
        db.query(Recommendation)
        .filter(Recommendation.transformer_id == transformer_id)
        .order_by(Recommendation.created_at.desc())
        .all()
    )

    pdf_bytes = _build_pdf(transformer, measurements, analyses, recommendations_list)
    safe_name = transformer.name.replace(" ", "_")

    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="FRA_Report_{safe_name}.pdf"'
        },
    )


# ── Excel Exports ───────────────────────────────────────────────────────────

def _build_measurements_excel(measurements, transformer_name: str | None = None) -> bytes:
    """Build an Excel workbook with measurement data."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Measurements"

    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(bottom=Side(style="thin", color="E2E8F0"))

    # Header row
    headers = [
        "Measurement ID", "Transformer", "Date", "Winding Config",
        "Vendor", "Temperature (°C)", "Freq Min (Hz)", "Freq Max (Hz)",
        "Data Points", "Created At",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows
    for row_idx, m in enumerate(measurements, 2):
        freq_min = min(m.frequency_hz) if m.frequency_hz else None
        freq_max = max(m.frequency_hz) if m.frequency_hz else None
        values = [
            m.id,
            getattr(m, "transformer_name", transformer_name or "—"),
            m.measurement_date.strftime("%Y-%m-%d") if m.measurement_date else None,
            m.winding_config,
            m.vendor,
            m.temperature_celsius,
            freq_min,
            freq_max,
            len(m.frequency_hz) if m.frequency_hz else 0,
            m.created_at.strftime("%Y-%m-%d %H:%M") if m.created_at else None,
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.border = thin_border

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_analyses_excel(analyses) -> bytes:
    """Build an Excel workbook with analysis results."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Analyses"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(bottom=Side(style="thin", color="E2E8F0"))

    headers = [
        "Analysis ID", "Measurement ID", "Date", "Fault Type",
        "Probability", "Confidence", "Health Score", "Model Version",
        "Axial Disp.", "Radial Def.", "Core Ground.", "Winding Short",
        "Loose Clamp.", "Moisture", "Healthy",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_idx, a in enumerate(analyses, 2):
        probs = a.all_probabilities or {}
        values = [
            a.id,
            a.measurement_id,
            a.created_at.strftime("%Y-%m-%d %H:%M") if a.created_at else None,
            (a.fault_type or "").replace("_", " ").title(),
            a.probability_score,
            a.confidence_level,
            a.health_score,
            a.model_version,
            probs.get("axial_displacement"),
            probs.get("radial_deformation"),
            probs.get("core_grounding"),
            probs.get("winding_short_circuit"),
            probs.get("loose_clamping"),
            probs.get("moisture_ingress"),
            probs.get("healthy"),
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.border = thin_border
            if isinstance(v, float):
                cell.number_format = "0.00%"

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/export/measurements")
def export_measurements_excel(
    transformer_id: str | None = Query(None, description="Filter by transformer"),
    db: Session = Depends(get_db),
):
    """Export measurements to an Excel file."""
    query = db.query(FRAMeasurement)
    filename_parts = ["FRA_Measurements"]

    if transformer_id:
        transformer = db.query(Transformer).filter(Transformer.id == transformer_id).first()
        if not transformer:
            raise HTTPException(404, "Transformer not found")
        query = query.filter(FRAMeasurement.transformer_id == transformer_id)
        filename_parts.append(transformer.name.replace(" ", "_"))

    measurements = query.order_by(FRAMeasurement.measurement_date.desc()).all()
    if not measurements:
        raise HTTPException(404, "No measurements found")

    excel_bytes = _build_measurements_excel(measurements)
    filename = "_".join(filename_parts) + ".xlsx"

    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export/analyses")
def export_analyses_excel(
    transformer_id: str | None = Query(None, description="Filter by transformer"),
    db: Session = Depends(get_db),
):
    """Export analysis results to an Excel file."""
    query = db.query(FaultAnalysis)
    filename_parts = ["FRA_Analyses"]

    if transformer_id:
        transformer = db.query(Transformer).filter(Transformer.id == transformer_id).first()
        if not transformer:
            raise HTTPException(404, "Transformer not found")
        measurement_ids = [
            m.id for m in
            db.query(FRAMeasurement.id)
            .filter(FRAMeasurement.transformer_id == transformer_id)
            .all()
        ]
        query = query.filter(FaultAnalysis.measurement_id.in_(measurement_ids))
        filename_parts.append(transformer.name.replace(" ", "_"))

    analyses = query.order_by(FaultAnalysis.created_at.desc()).all()
    if not analyses:
        raise HTTPException(404, "No analyses found")

    excel_bytes = _build_analyses_excel(analyses)
    filename = "_".join(filename_parts) + ".xlsx"

    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
